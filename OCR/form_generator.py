from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension, DimensionHolder
from openpyxl.utils import get_column_letter

import json


with open("forms/sample.json","r") as json_obj:
    # print(json_obj)
    page_format = json.load(json_obj)
    
    page_header = page_format["page_header"]
    form_format = page_format["form_format"]

print(page_header)
print(form_format)
no_of_questions = len(form_format)
print("there are {} questions".format(no_of_questions))


wb = Workbook()

# grab the active worksheet
ws = wb.active

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True
ws.oddHeader.left.text = page_header

print(ws.HeaderFooter)

for i in range(0,11):
    ws.column_dimensions[get_column_letter(i+1)].width = 11

border = Border(left=Side(border_style=None,
                            color='000000',
                            style='thin'),
                right=Side(border_style=None,
                            color='000000',
                            style='thin'),
                top=Side(border_style=None,
                            color='000000',
                            style='thin'),
                bottom=Side(border_style=None,
                            color='000000',
                            style='thin'))

for column in ws['A1:K32']:
    for cell in column:
        cell.border = border

ws.cell(row=1,column=1).value='casdfa'

for col in range(0,11):

    if col+1 > no_of_questions:
        break

    for row in range(0, 32):

        value = 'x'

        if row == 0:
        
            value = form_format[str(col+1)]["qheader"]
        
        else:

            if form_format[str(col+1)]["qtype"] == 1:

                value = ''

                for i, choices in enumerate(form_format[str(col+1)]["choices"]):
                    if i == len(form_format[str(col+1)]["choices"]):
                        value = value+'O'
                    else:
                        value = value+'O '

            elif form_format[str(col+1)]["qtype"] == 2:

                value = ''

        ws.cell(row=row+1,column=col+1).value = value

# Save the file
wb.save("forms/out/sample.xlsx")